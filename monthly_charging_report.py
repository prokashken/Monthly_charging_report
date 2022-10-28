import os
import datetime
import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import numbers
from glob import glob



# process of Merging of all downloaded file into merged_file


#alternative to append
stock_files = sorted(glob('file_*.csv'))

master = pd.concat((pd.read_csv(file)
	for file in stock_files), ignore_index = True)

master.to_csv('merged_file.csv', index = False)




'''
master_df = pd.DataFrame()

for file in os.listdir(os.getcwd()):
	if file.endswith('.csv'):
		master_df = master_df.append(pd.read_csv(file))

master_df.to_csv('merged_file.csv', index=False)
'''

# modifying the formate of start_time & stop_time  and put the value into modyfy.csv & Book2_cp.xlsx

readfile = pd.read_csv("merged_file.csv")

readfile1 = readfile[['site_id','charger_id','start_time','stop_time']]
readfile1['start_time'] = pd.to_datetime(readfile1.start_time)
readfile1['stop_time'] = pd.to_datetime(readfile1.stop_time)

df = readfile1.sort_values(by=['site_id'], ascending=False)
df.reset_index(drop=True, inplace=True)

#df['Duration'] = df['stop_time'] - df['start_time']
#readfile.to_excel("Book2.xlsx")


#f = datetime.time(df['start_time'])
#print(f)

#print(df['Duration'])
df.to_csv("modify.csv")

readfile2 = pd.read_csv("modify.csv")

readfile2.to_excel("Book2_cp.xlsx")


wb = openpyxl.load_workbook("Book2_cp.xlsx")
ws = wb['Sheet1']
m = 2
n = 2
rows = ws.max_row


#Removing "+00:00" from start_time(E) & stop_time(F) column 

#ws['E2'].number_format = numbers.FORMAT_DATE_DATETIME
#ws['F2'].number_format = numbers.FORMAT_DATE_DATETIME
rows = ws.max_row
for i in range(2,rows+1):
	date1 = ws['E'+str(i)].value
	date2 = ws['F'+str(i)].value
	whatWeNeed = date1.replace("+00:00","")
	whatWeNeed2 = date2.replace("+00:00","")
	ws['E'+str(i)] = whatWeNeed
	ws['F'+str(i)] = whatWeNeed2

wb.save("Book2_cp.xlsx")



# Calculation Duration on column = 7(G)

for i in range(2,rows+1):
	#start_time = ws.cell(row=i,column=4).value
	#end_time = ws.cell(row=i,column=4).value
	#duration = (end_time - start_time)
	duration = "=(F"+str(i)+"-"+"E"+str(i)
	ws.cell(row=i,column=7).value = duration
	#ws['E'+str(i)] = duration

wb.save("Book2_cp.xlsx")



#Calcultatin total duration and mearge the cells

#ws['F2'] ="=SUM(E2:E17)"
#ws.merge_cells('F2:F17')
for i in range(2,rows+1):  # 2:8[value 1  baraiya rakte hobe row thika]
	c = ws["C"+str(i)]
	#print(i)
	#print(m)
	#print(ws['A'+str(m)].value)
	#print(c.value)
	#print(c.value != ws['A'+str(m)].value)
	if c.value != ws['C'+str(m)].value:
		start = "G"+str(m)
		end = "G"+str(i-1)
		#ws['H'+str(m)] = "=SUM("+start+":"+end+")"
		ws.merge_cells("H"+str(m)+":"+"H"+str(i-1))
		ws['H'+str(m)] = "=SUM("+start+":"+end+")"
		m = i

	if i == (rows):
		start = "G"+str(m)
		end = "G"+str(i)
		#ws['H'+str(m)] = "=SUM("+start+":"+end+")"
		ws.merge_cells("H"+str(m)+":"+"H"+str(i))
		ws['H'+str(m)] = "=SUM("+start+":"+end+")"
		m = i
	#print(c.value)


#wb.save("Book2.xlsx")

wb.save("Book2_cp.xlsx")


#############Putting Site ID In Sheet

Des_data = openpyxl.load_workbook('Book2_cp.xlsx')
sor_data = openpyxl.load_workbook('from.xlsx')

Des_sheet = Des_data['Sheet1']
sor_sheet = sor_data['Sheet1']

for p in Des_sheet.iter_rows():
	id = p[2].value
	row_number = p[2].row
	for q in sor_sheet.iter_rows():
		if q[1].value == id:
			Des_sheet.cell(row=row_number, column=2).value = q[0].value
			#print(q[0].value)



for z in range(2,rows+1):  # 2:8[value 1  baraiya rakte hobe row thika]
	c = Des_sheet["C"+str(z)]
	#print(c.value)
	#print("n = ",Des_sheet['C'+str(n)].value)
	if c.value != Des_sheet['C'+str(n)].value:
		#start = "G"+str(m)
		#end = "G"+str(i-1)
		#ws['H'+str(m)] = "=SUM("+start+":"+end+")"
		Des_sheet.merge_cells("B"+str(n)+":"+"B"+str(z-1))
		#ws['H'+str(m)] = "=SUM("+start+":"+end+")"
		n = z

	if z == (rows):
		#start = "G"+str(n)
		#end = "G"+str(i)
		#ws['H'+str(m)] = "=SUM("+start+":"+end+")"
		Des_sheet.merge_cells("B"+str(n)+":"+"B"+str(z))
		#ws['H'+str(m)] = "=SUM("+start+":"+end+")"
		n = z

Des_sheet['G1'] = "Duration"
Des_sheet['H1'] = "Total Duration"
Des_sheet['B1'] = "Site_name"

Des_data.save('Book2_cp.xlsx')	


